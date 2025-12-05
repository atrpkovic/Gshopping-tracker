#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Google Shopping brand-hit tracker (parallel, multi-sheet in/out, multi-location).

APPEND-ONLY OUTPUTS:
- XLSX: appends rows to each sheet in OUTPUT_XLSX; preserves existing headers/rows.
- CSV:  appends rows to per-sheet CSV files in CSV_FALLBACK_DIR; header only on first write.

Notes:
- Loads SERPAPI_KEY from env (or .env if python-dotenv is installed)
- Reads keywords from Excel/CSV. If Excel with multiple sheets: each sheet is processed separately.
- Queries SerpAPI "google_shopping" for each keyword × each location (in parallel)
- Matches results by brand synonyms (in-code defaults; optional JSON override via BRANDS_JSON_PATH)
"""

from __future__ import annotations

import os
import re
import sys
import json
import time
import random
import logging
import pathlib
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict, OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

# --- Optional .env loader (safe if not installed) ---
try:
    from pathlib import Path
    from dotenv import load_dotenv  # type: ignore
    load_dotenv(dotenv_path=Path(__file__).with_name(".env"))
    load_dotenv()
except Exception:
    pass

import requests
import pandas as pd

# ----------------------------
# Configuration
# ----------------------------
KEYWORDS_FILE = os.environ.get("KEYWORDS_FILE", "keywords.xlsx")  # set to "keywords_test.xlsx" while testing
SHEETS = [s.strip() for s in os.environ.get("KEYWORDS_SHEETS", "").split(",") if s.strip()]

OUTPUT_XLSX = os.environ.get("OUTPUT_XLSX", "google_shopping_brand_hits.xlsx")
CSV_FALLBACK_DIR = os.environ.get("CSV_FALLBACK_DIR", "brand_hits_csv")

SERPAPI_KEY = os.getenv("SERPAPI_KEY")
SERP_ENGINE = "google_shopping"
GOOGLE_DOMAIN = os.environ.get("GOOGLE_DOMAIN", "google.com")
GL = os.environ.get("GL", "us")
HL = os.environ.get("HL", "en")

# Speed / politeness
MAX_WORKERS = int(os.environ.get("MAX_WORKERS", "6"))
SERPAPI_QPS = float(os.environ.get("SERPAPI_QPS", "2"))  # global QPS cap across all workers
MAX_RESULTS = int(os.environ.get("MAX_RESULTS", "50"))
MAX_API_RETRIES = int(os.environ.get("MAX_API_RETRIES", "2"))
RETRY_BACKOFF_BASE = float(os.environ.get("RETRY_BACKOFF_BASE", "1.6"))

BRANDS_JSON_PATH = os.environ.get("BRANDS_JSON_PATH", "").strip() or None

# Default brand map (override with BRANDS_JSON_PATH if provided)
DEFAULT_BRANDS_MAP: Dict[str, List[str]] = {
    "prioritytire.com": ["priority tire", "prioritytire"],
    "simpletire.com":   ["simple tire", "simpletire"],
    "tireagent.com":    ["tire agent", "tireagent"],
    "gigatires.com":    ["giga tires", "gigatires"],
    "tireseasy.com":    ["tireseasy", "tires easy"],
    "tirerack.com":     ["tirerack", "tire rack"],
    "ebay.com":         ["ebay"],
    "walmart.com":      ["walmart"],
}

# Locations (label, SerpAPI "location" input)
LOCATIONS: List[Tuple[str, str]] = [
    ("Florida, United States",        "Florida, United States"),
    ("Texas, United States",          "Texas, United States"),
    ("North Carolina, United States", "North Carolina, United States"),
    ("Pennsylvania, United States",   "Pennsylvania, United States"),
    ("California, United States",     "California, United States"),
    ("Illinois, United States",       "Illinois, United States"),
]

LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()

COLS = [
    "Timestamp",
    "Sheet",
    "Keyword",
    "Location",
    "Position",
    "Title",
    "Price",
    "Price Value",
    "Old Price",
    "Old Price Value",
    "Rating",
    "Reviews",
    "Merchant Name",
    "Product Link",
    "Brand",
    "Matched Synonym",
]

# ------------- Logging -------------
logger = logging.getLogger("tracker")
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s"))
logger.addHandler(_handler)
from pathlib import Path

LOG_FILE = Path(__file__).with_suffix(".log")

file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s"))
logger.addHandler(file_handler)

# ------------- Helpers -------------

def _now() -> str:
    return pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S")

def _is_excel(fname: str) -> bool:
    low = fname.lower()
    return low.endswith(".xlsx") or low.endswith(".xlsm") or low.endswith(".xls")

def _is_csv(fname: str) -> bool:
    return fname.lower().endswith(".csv")

def _normalize_price_str(p: Optional[str]) -> Tuple[str, Optional[float]]:
    if not p:
        return "", None
    s = str(p)
    raw = s
    s2 = re.sub(r"[^0-9.,]", "", s)
    try:
        if s2.count(",") and s2.count("."):
            s3 = s2.replace(",", "")
        else:
            if s2.count(",") == 1 and s2.count(".") == 0:
                s3 = s2.replace(",", ".")
            else:
                s3 = s2.replace(",", "")
        return raw, float(s3)
    except Exception:
        return raw, None

def _read_keywords_multi_sheet(path: str, only_sheets: Optional[List[str]] = None) -> Dict[str, List[str]]:
    store: Dict[str, List[str]] = {}
    if _is_excel(path):
        xl = pd.ExcelFile(path)
        candidate_sheets = only_sheets or xl.sheet_names
        for sh in candidate_sheets:
            df = pd.read_excel(path, sheet_name=sh)
            cols = [c.strip().lower() for c in df.columns]
            if "keyword" not in cols:
                logger.warning(f"Sheet '{sh}' has no 'keyword' column; skipping.")
                continue
            kw_col = df.columns[cols.index("keyword")]
            kws = [str(x).strip() for x in df[kw_col].dropna().astype(str).tolist() if str(x).strip()]
            if not kws:
                logger.info(f"Sheet '{sh}' contains 0 keywords; skipping.")
                continue
            store[sh] = kws
    elif _is_csv(path):
        df = pd.read_csv(path)
        cols = [c.strip().lower() for c in df.columns]
        if "keyword" not in cols:
            raise ValueError("CSV has no 'keyword' column.")
        kw_col = df.columns[cols.index("keyword")]
        kws = [str(x).strip() for x in df[kw_col].dropna().astype(str).tolist() if str(x).strip()]
        store["Sheet1"] = kws
    else:
        raise FileNotFoundError(f"Unsupported keywords file: {path}")
    total = sum(len(v) for v in store.values())
    logger.info(f"Loaded {total} keywords across {len(store)} sheet(s)")
    return store

def load_brands_map(brands_json_path: Optional[str]) -> Dict[str, List[str]]:
    if brands_json_path:
        p = pathlib.Path(brands_json_path)
        if p.is_file():
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                if isinstance(data, dict) and all(isinstance(v, list) for v in data.values()):
                    return data
                logger.warning(f"Invalid brands JSON at '{brands_json_path}'; using in-code defaults.")
            except Exception as e:
                logger.warning(f"Failed reading brands JSON '{brands_json_path}': {e}; using in-code defaults.")
    return DEFAULT_BRANDS_MAP

def compile_brand_matchers(brands_map: Dict[str, List[str]]) -> List[Tuple[str, re.Pattern]]:
    compiled: List[Tuple[str, re.Pattern]] = []
    for domain, synonyms in brands_map.items():
        esc = [re.escape(s.strip()) for s in synonyms if s.strip()]
        if not esc:
            continue
        pattern = r"(?i)\b(" + "|".join(esc) + r")\b"
        compiled.append((domain, re.compile(pattern)))
    return compiled

# ------------- Polite global rate limiter -------------

class GlobalRateLimiter:
    """
    Very simple global QPS limiter shared by all threads.
    Ensures we don't exceed SERPAPI_QPS across the whole process.
    """
    def __init__(self, qps: float):
        self.min_interval = 1.0 / max(0.001, qps)
        self._lock = Lock()
        self._last = 0.0

    def wait(self):
        with self._lock:
            now = time.time()
            elapsed = now - self._last
            if elapsed < self.min_interval:
                sleep_s = self.min_interval - elapsed + random.uniform(0.0, 0.05)
                time.sleep(sleep_s)
                self._last = time.time()
            else:
                self._last = now

rate_limiter = GlobalRateLimiter(SERPAPI_QPS)

# ------------- SerpAPI -------------

class SerpApiError(Exception):
    pass

def serp_shopping_search(
    query: str,
    location: str,
    api_key: str,
    num: int = 50,
    gl: str = "us",
    hl: str = "en",
    google_domain: str = "google.com",
    retries: int = 2,
    backoff_base: float = 1.6,
) -> Dict[str, Any]:
    url = "https://serpapi.com/search.json"
    params = {
        "engine": SERP_ENGINE,
        "q": query,
        "location": location,
        "google_domain": google_domain,
        "gl": gl,
        "hl": hl,
        "num": max(10, min(int(num), 100)),
        "api_key": api_key,
    }

    for attempt in range(1, max(1, retries) + 1):
        # global QPS gate
        rate_limiter.wait()

        try:
            r = requests.get(url, params=params, timeout=60)
            if r.status_code == 200:
                return r.json()
            if r.status_code in (429, 500, 502, 503, 504):
                delay = (backoff_base ** (attempt - 1)) + random.uniform(0, 1.2)
                logger.warning(f"SerpAPI HTTP {r.status_code} (attempt {attempt}/{retries}); sleeping {delay:.1f}s …")
                time.sleep(delay)
                continue
            raise SerpApiError(f"SerpAPI HTTP {r.status_code}: {r.text[:240]}")
        except requests.RequestException as e:
            delay = (backoff_base ** (attempt - 1)) + random.uniform(0, 1.2)
            logger.warning(f"SerpAPI error {e!r} (attempt {attempt}/{retries}); sleeping {delay:.1f}s …")
            time.sleep(delay)

    raise SerpApiError("SerpAPI request failed after retries.")

# ------------- Parsing & matching -------------

def extract_shopping_results(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    if not payload:
        return results
    items = payload.get("shopping_results") or []
    for idx, it in enumerate(items, 1):
        title = it.get("title") or ""
        link = it.get("link") or it.get("product_link") or ""
        merchant = it.get("source") or it.get("store") or it.get("seller") or ""
        price_str = it.get("price") or it.get("extracted_price") or ""
        extracted_price = it.get("extracted_price")
        if extracted_price is None:
            _, price_val = _normalize_price_str(price_str if isinstance(price_str, str) else str(price_str))
        else:
            price_val = float(extracted_price)

        old_price_str = it.get("original_price") or it.get("old_price") or ""
        old_price_val = None
        if old_price_str:
            _, old_price_val = _normalize_price_str(str(old_price_str))

        rating = it.get("rating")
        reviews = it.get("reviews")

        results.append({
            "Position": idx,
            "Title": title,
            "Product Link": link,
            "Merchant Name": merchant,
            "Price": price_str if isinstance(price_str, str) else (f"${price_val:.2f}" if price_val else ""),
            "Price Value": price_val,
            "Old Price": old_price_str if isinstance(old_price_str, str) else (f"${old_price_val:.2f}" if old_price_val else ""),
            "Old Price Value": old_price_val,
            "Rating": rating if isinstance(rating, (int, float)) else None,
            "Reviews": reviews if isinstance(reviews, (int, float)) else None,
        })
    return results

def match_brand(row: Dict[str, Any], brand_patterns: List[Tuple[str, re.Pattern]]) -> Tuple[Optional[str], Optional[str]]:
    hay = " ".join([
        str(row.get("Title", "") or ""),
        str(row.get("Merchant Name", "") or "")
    ]).lower()
    for domain, creg in brand_patterns:
        m = creg.search(hay)
        if m:
            return domain, m.group(1)
    return None, None

# ------------- Append-ONLY writing -------------

def append_xlsx_per_sheet(out_path: str, sheet2df: Dict[str, pd.DataFrame]) -> None:
    """
    Appends rows to an XLSX book per sheet, preserving headers.
    - If file doesn't exist: create and write headers + rows.
    - If sheet doesn't exist: create with headers + rows.
    - If sheet exists: append rows under existing header order.
    """
    from openpyxl import load_workbook, Workbook

    p = pathlib.Path(out_path)
    if not p.exists():
        # Create new workbook with all sheets + headers in one go
        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)
        for sheet, df in sheet2df.items():
            ws = wb.create_sheet(title=(sheet or "Sheet1")[:31])
            # Ensure full column set in defined order
            df2 = df.reindex(columns=COLS, fill_value="")
            # Write header
            ws.append(COLS)
            # Write rows
            for row in df2.itertuples(index=False, name=None):
                ws.append(row)
        wb.save(out_path)
        return

    # File exists -> open and append
    wb = load_workbook(out_path)
    for sheet, df in sheet2df.items():
        # Nothing to append
        if df is None or df.empty:
            # Ensure sheet exists with header at least
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(title=(sheet or "Sheet1")[:31])
                ws.append(COLS)
            continue

        safe_sheet = (sheet or "Sheet1")[:31]
        if safe_sheet in wb.sheetnames:
            ws = wb[safe_sheet]
            # Ensure header row exists; if empty sheet, write header
            if ws.max_row < 1 or all(c.value is None for c in ws[1]):
                ws.append(COLS)
                existing_headers = COLS
            else:
                existing_headers = [cell.value for cell in ws[1]]
                # If headers are missing/partial, normalize to COLS
                if not existing_headers or any(h is None for h in existing_headers):
                    existing_headers = COLS
            # Align df columns to existing header order; add blanks for missing
            df2 = df.copy()
            for h in existing_headers:
                if h not in df2.columns:
                    df2[h] = ""
            df2 = df2[existing_headers]
            for row in df2.itertuples(index=False, name=None):
                ws.append(row)
        else:
            ws = wb.create_sheet(title=safe_sheet)
            ws.append(COLS)
            df2 = df.reindex(columns=COLS, fill_value="")
            for row in df2.itertuples(index=False, name=None):
                ws.append(row)

    wb.save(out_path)

def append_csv_folder(base_dir: str, sheet2df: Dict[str, pd.DataFrame]) -> None:
    """
    Appends rows to CSV per sheet. Creates file + header if missing.
    """
    base = pathlib.Path(base_dir)
    base.mkdir(parents=True, exist_ok=True)

    for sheet, df in sheet2df.items():
        if df is None or df.empty:
            continue
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", sheet)[:64] or "Sheet1"
        p = base / f"{safe}.csv"

        df2 = df.reindex(columns=COLS, fill_value="")
        write_header = not p.exists()
        mode = "a" if p.exists() else "w"
        df2.to_csv(p, index=False, encoding="utf-8-sig", mode=mode, header=write_header)

# ------------- Task & worker -------------

def hashlib_stub(s: str) -> str:
    try:
        import hashlib
        return hashlib.md5(s.encode("utf-8")).hexdigest()[:24]
    except Exception:
        return "location"

def _build_tasks(sheets_kws: Dict[str, List[str]]) -> List[Tuple[str, str, str, str]]:
    """
    Returns list of (sheet_name, keyword, loc_label, loc_input) tasks.
    """
    tasks: List[Tuple[str, str, str, str]] = []
    for sheet, kws in sheets_kws.items():
        for kw in kws:
            for (loc_label, loc_string) in LOCATIONS:
                tasks.append((sheet, kw, loc_label, loc_string))
    return tasks

def _worker_task(
    sheet: str,
    kw: str,
    loc_label: str,
    loc_string: str,
    api_key: str,
    brand_patterns: List[Tuple[str, re.Pattern]],
) -> List[Dict[str, Any]]:
    """
    Executes one keyword × location call, returns rows (already filtered by brand).
    """
    rows: List[Dict[str, Any]] = []
    try:
        payload = serp_shopping_search(
            query=kw,
            location=loc_string,
            api_key=api_key,
            num=MAX_RESULTS,
            gl=GL,
            hl=HL,
            google_domain=GOOGLE_DOMAIN,
            retries=MAX_API_RETRIES,
            backoff_base=RETRY_BACKOFF_BASE,
        )
    except SerpApiError as e:
        logger.error(f"[{sheet}] '{kw}' @ {loc_label}: {e}")
        return rows

    items = extract_shopping_results(payload)
    if not items:
        return rows

    for it in items:
        domain, syn = match_brand(it, brand_patterns)
        if not domain:
            continue
        
        # ---------------------------------------------------------
        # OVERRIDE: Extract specific seller from eBay/Walmart
        # e.g. "eBay - unitedtires" -> Brand="unitedtires"
        # ---------------------------------------------------------
        merchant = it.get("Merchant Name")
        if merchant and " - " in str(merchant):
            parts = str(merchant).split(" - ", 1)
            prefix = parts[0].strip().lower()
            # If the domain (marketplace) matched is ebay or walmart, assume
            # the part after the hyphen is the actual seller we want to track.
            if "ebay" in prefix or "walmart" in prefix:
                real_seller = parts[1].strip()
                domain = real_seller        # Override Brand (Column O)
                syn = real_seller           # Override Synonym (Column P)
        # ---------------------------------------------------------

        rows.append({
            "Timestamp": _now(),
            "Sheet": sheet,
            "Keyword": kw,
            "Location": loc_label,
            "Position": it.get("Position"),
            "Title": it.get("Title"),
            "Price": it.get("Price"),
            "Price Value": it.get("Price Value"),
            "Old Price": it.get("Old Price"),
            "Old Price Value": it.get("Old Price Value"),
            "Rating": it.get("Rating"),
            "Reviews": it.get("Reviews"),
            "Merchant Name": it.get("Merchant Name"),
            "Product Link": it.get("Product Link"),
            "Brand": domain,
            "Matched Synonym": syn,
        })
    return rows

# ------------- Main -------------

def run() -> None:
    if not SERPAPI_KEY:
        logger.error("Missing SERPAPI_KEY. Set it in your environment or .env file.")
        return

    if not pathlib.Path(KEYWORDS_FILE).is_file():
        logger.error(f"Keywords file not found: {KEYWORDS_FILE}")
        return

    brands_map = load_brands_map(BRANDS_JSON_PATH)
    brand_patterns = compile_brand_matchers(brands_map)

    sheets_kws = _read_keywords_multi_sheet(KEYWORDS_FILE, SHEETS if SHEETS else None)

    logger.info(f"📍 Multi-location mode: {len(LOCATIONS)} locations")
    for label, code in LOCATIONS:
        logger.info(f"   - {label} (code: {hashlib_stub(code)})")

    total_keywords = sum(len(v) for v in sheets_kws.values())
    logger.info(f"Loaded {total_keywords} keywords across {len(sheets_kws)} sheet(s)")

    tasks = _build_tasks(sheets_kws)
    logger.info(f"Scheduling {len(tasks)} keyword×location calls with MAX_WORKERS={MAX_WORKERS}, SERPAPI_QPS={SERPAPI_QPS}")

    # Collect rows per sheet
    per_sheet_rows: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    with ThreadPoolExecutor(max_workers=max(1, MAX_WORKERS)) as ex:
        futures = {
            ex.submit(_worker_task, sheet, kw, loc_label, loc_string, SERPAPI_KEY, brand_patterns): (sheet, kw, loc_label)
            for (sheet, kw, loc_label, loc_string) in tasks
        }
        for fut in as_completed(futures):
            sheet, kw, loc_label = futures[fut]
            try:
                rows = fut.result()
                if rows:
                    per_sheet_rows[sheet].extend(rows)
            except Exception as e:
                logger.error(f"Worker failed for [{sheet}] '{kw}' @ {loc_label}: {e}")

    # Build frames per sheet in input order
    out_frames: "OrderedDict[str, pd.DataFrame]" = OrderedDict()
    for sheet in sheets_kws.keys():
        rows = per_sheet_rows.get(sheet, [])
        df_sheet = pd.DataFrame(rows, columns=COLS) if rows else pd.DataFrame(columns=COLS)
        out_frames[sheet] = df_sheet

    # ---- APPEND to BOTH outputs every time ----
    try:
        append_xlsx_per_sheet(OUTPUT_XLSX, out_frames)
        logger.info(f"XLSX appended: {OUTPUT_XLSX}")
    except PermissionError as e:
        logger.error(f"Failed to append XLSX (likely open/locked): {e}")

    append_csv_folder(CSV_FALLBACK_DIR, out_frames)
    logger.info(f"Per-sheet CSVs appended in: {CSV_FALLBACK_DIR}")

    if all(df.empty for df in out_frames.values()):
        logger.info("No rows collected.")

if __name__ == "__main__":
    run()